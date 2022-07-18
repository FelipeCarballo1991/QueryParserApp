import mysql.connector
import psycopg2
from conexion import conexionDW
import openpyxl
from openpyxl.styles import Font
from query_formatter import query_columns

def export_to_excel(connection, query_string, headings, filepath = "matricula.xlsx"):

    """
    Exports data from PostgreSQL to an Excel spreadsheet using psycopg2.

    Arguments:
    connection - an open psycopg2 (this function does not close the connection)
    query_string - SQL to get data
    headings - list of strings to use as column headings
    filepath - path and filename of the Excel file

    psycopg2 and file handling errors bubble up to calling code.
    """

    cursor = connection.cursor()
    cursor.execute(query_string)
    data = cursor.fetchall()
    cursor.close()

    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.row_dimensions[1].font = Font(bold = True)

    # Spreadsheet row and column indexes start at 1
    # so we use "start = 1" in enumerate so
    # we don't need to add 1 to the indexes.
    for colno, heading in enumerate(headings, start = 1):
        sheet.cell(row = 1, column = colno).value = heading

    # This time we use "start = 2" to skip the heading row.
    for rowno, row in enumerate(data, start = 2):
        for colno, cell_value in enumerate(row, start = 1):
            sheet.cell(row = rowno, column = colno).value = cell_value

    wb.save(filepath)
    connection.close()


def connection (host,database,user,password):    
   
    connection = psycopg2.connect(host=host, 
                                database=database, 
                                user=user, 
                                password=password)

    return connection

if __name__ == "__main__":
    
    try:
        conexion = connection(conexionDW["host"],conexionDW["database"],conexionDW["user"],conexionDW["password"])
        print("Conexion a Postgres")
    except:
        print("Error de conexión")

query_string = "select cue, descripcion_nivel, descripcion, de_, df_descripcion, apellido, nombre, documento, descripcion_anio, division, descripcion_turno, id_localizacion, fecha_upd_alumno_movimiento, tipo_documento, desc_plan_estudio, ax, modalidad, descripcion_pais, genero, fecha_nacimiento, cl, estado, origen, id_seccion, id_plan_estudio, id_alumno, email, id_usuario, username, pass, jornada, td_ab from miescuelamg.matricula m limit 1"
columnas = query_columns(query_string[:-1])

export_to_excel(conexion, query_string, columnas)

conexion.close()
